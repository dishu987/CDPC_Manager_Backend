from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from Company.serializers.register import CompanyAddSerializer,CompanyGETSerializer,CompanyJobTagSerializer,CompanyHiringTagSerializer
from Company.models import Company,CompanyJobTag,CompanyHiringTag,Feedback,HRDetails
from rest_framework.permissions import IsAuthenticated
from users.utils.rederers.user import UserRenderer
from users.permissions.roles import IsHeadPlacementCoordinator,IsSuperUser,IsStudentCoordinator,IsManager
from django.shortcuts import get_object_or_404
from .serializers.comments import CommentSerializer,CommentAddSerializer
from .models import Comment
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.http import HttpResponse



class CompanyAddView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated, IsSuperUser]
    def post(self, request, format=None):
        serializer = CompanyAddSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        company = serializer.save()
        return Response({'msg': 'Company added successfully'}, status=status.HTTP_201_CREATED)


class CompanyListView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    def get(self, request, format=None):
        companies = Company.objects.all()
        serializer = CompanyGETSerializer(companies, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class CompanyFilterAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    def get(self, request):
        # Get the filter parameters from the request data
        request_data = request.query_params
        active = request_data.get("active", None)
        importance_start = request_data.get("importance__start", None)
        importance_end = request_data.get("importance__end", None)
        salary_start = request_data.get("salary__start", None)
        salary_end = request_data.get("salary__end", None)
        hiring_tags = request_data.getlist("hiring_tags", None)
        job_tags = request_data.getlist("job_tags", None)

        # Create the filter conditions based on the provided parameters
        filters = {}
        if active is not None:
            filters["status"] = active
        if importance_start is not None and importance_end is not None:
            filters["importance__range"] = (importance_start, importance_end)
        if salary_start is not None and salary_end is not None:
            filters["salary__range"] = (salary_start, salary_end)
        if hiring_tags:
            filters["hiring_tags__in"] = hiring_tags
        if job_tags:
            filters["job_tags__in"] = job_tags

        # Retrieve the filtered companies
        try:
            companies = Company.objects.filter(**filters)
            serializer = CompanyGETSerializer(companies, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({"error": "Something went wrong.."}, status=status.HTTP_400_BAD_REQUEST)



class CompanyJobTagsView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        tags = CompanyJobTag.objects.all()
        serializer = CompanyJobTagSerializer(tags, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class CompanyHiringTagsView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        tags = CompanyHiringTag.objects.all()
        serializer = CompanyHiringTagSerializer(tags, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class CompanyDetailsView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        company_id = request.query_params.get('id')
        company = get_object_or_404(Company, id=company_id)

        job_tags = []
        hiring_tags = []
        hrs = []
        assigned_coordinators = []
        spoc = None

        for job_tag in company.job_tags.all():
            job_tags.append({
                "id": job_tag.pk,
                "name": job_tag.name,
            })
        for hiring_tag in company.hiring_tags.all():
            hiring_tags.append({
                "id": hiring_tag.pk,
                "name": hiring_tag.name,
            })
        for hr in company.hr_details.all():
            hrs.append({
                "id": hr.pk,
                "name": hr.name,
                "email": hr.email,
                "phone": hr.phone_number,
            })

        for coordinator in company.assigned_coordinators.all():
            assigned_coordinators.append({
                "id": coordinator.id,
                "name": coordinator.name,
                "email": coordinator.email,
            })

        if company.spoc:
            spoc = {
                "id": company.spoc.id,
                "name": company.spoc.name,
                "email": company.spoc.email,
            }

        data = {
            "id": company.id,
            "name": company.name,
            "about": company.about,
            "assigned_coordinators": assigned_coordinators,
            "salary": str(company.salary),
            "importance": company.importance,
            "years_of_collaboration": company.years_of_collaboration,
            "spoc": spoc,
            "job_location": company.job_location,
            "blacklist": company.blacklist,
            "status": company.status,
            "job_tags": job_tags,
            "hiring_tags": hiring_tags,
            "hr_details": hrs
        }
        return Response(data, status=status.HTTP_200_OK)



class CommentsListView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        company_id = request.query_params.get('id')
        try:
            comments = Comment.objects.filter(company_id=company_id, reply_to__isnull=True).order_by('-created_at')
            serializer = CommentSerializer(comments, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Comment.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)



class AddCommentView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]
    def post(self, request, format=None):
        serializer = CommentAddSerializer(data=request.data,context={'request': request})
        if serializer.is_valid():
            comment = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CommentDeleteAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated, IsSuperUser]
    def delete(self, request, comment_id):
        user = request.user
        try:
            comment = Comment.objects.get(id=comment_id)
        except Comment.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        comment.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CompanyStatusAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated, IsSuperUser]
    def put(self, request, pk):
        try:
            company = Company.objects.get(pk=pk)
        except Company.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        company.status = True
        company.save()
        return Response(status=status.HTTP_200_OK)

class CompanyBlacklistRemoveAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated, IsSuperUser]
    def put(self, request, pk):
        try:
            company = Company.objects.get(pk=pk)
        except Company.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        company.blacklist = False
        company.save()
        return Response(status=status.HTTP_200_OK)


class CompanyBlacklistAddAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated, IsSuperUser]
    def put(self, request, pk):
        try:
            company = Company.objects.get(pk=pk)
        except Company.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        company.blacklist = True
        company.save()
        return Response(status=status.HTTP_200_OK)



class DownloadCompaniesAPIView(APIView):
    renderer_classes = [UserRenderer]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Retrieve all companies from the database
        companies = Company.objects.all()

        # Create a new workbook
        workbook = Workbook()
        sheet = workbook.active

        # Set the column headers
        headers = [
            'Name', 'About', 'Assigned Coordinators', 'Salary', 'Importance', 'Years of Collaboration',
            'SPOC', 'Job Location', 'Blacklist', 'Job Tags','Hiring Tags', 'Status', 'HR Name', 'HR Email', 'HR Phone Number'
        ]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write the data rows
        row_number = 2
        for company in companies:
            assigned_coordinators = ', '.join(company.assigned_coordinators.values_list('email', flat=True))
            job_tags = ', '.join(company.job_tags.values_list('name', flat=True))
            hiring_tags = ', '.join(company.hiring_tags.values_list('name', flat=True))

            hr_details = company.hr_details.all()
            for hr in hr_details:
                sheet.cell(row=row_number, column=1, value=company.name)
                sheet.cell(row=row_number, column=2, value=company.about)
                sheet.cell(row=row_number, column=3, value=assigned_coordinators)
                sheet.cell(row=row_number, column=4, value=str(company.salary))
                sheet.cell(row=row_number, column=5, value=str(company.importance))
                sheet.cell(row=row_number, column=6, value=str(company.years_of_collaboration))
                sheet.cell(row=row_number, column=7, value=company.spoc.email if company.spoc else '')
                sheet.cell(row=row_number, column=8, value=company.job_location)
                sheet.cell(row=row_number, column=9, value=str(company.blacklist))
                sheet.cell(row=row_number, column=10, value=job_tags)
                sheet.cell(row=row_number, column=11, value=hiring_tags)
                sheet.cell(row=row_number, column=12, value=str(company.status))
                sheet.cell(row=row_number, column=13, value=hr.name)
                sheet.cell(row=row_number, column=14, value=hr.email)
                sheet.cell(row=row_number, column=15, value=hr.phone_number)

                row_number += 1

        # Save the workbook to a BytesIO object
        file_stream = io.BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        # Create the response with appropriate headers
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="companies.xlsx"'
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # Set the file content from the BytesIO object
        response.write(file_stream.getvalue())

        return response
